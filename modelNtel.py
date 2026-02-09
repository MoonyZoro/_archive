from matplotlib import animation
import numpy as np
import sys
import os
import json
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout,
QHBoxLayout,
QPushButton, QListWidget, QWidget, QLabel, QLineEdit,
QListWidgetItem, QDialog, QFormLayout, QSpinBox,
QDoubleSpinBox, QMessageBox, QScrollArea,
QInputDialog, QMenu, QAction, QComboBox, QTextEdit, QSplitter, QGroupBox,
QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog)
from PyQt5.QtCore import Qt
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
from mpl_toolkits.mplot3d import Axes3D

class settings():
    dt = None
    def set_data(self, dt):
        self.dt = dt

global se
se = settings()
se.dt = 24*60*60

# Глобальные переменные
SELECTED_SAVE = None
global all_data
all_data = []

def animation3dyes(all_data):
    # Импортируем необходимые модули для интеграции с PyQt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel
    from PyQt5.QtCore import QTimer

    class AnimationDialog(QDialog):
        def __init__(self, all_data):
            super().__init__()
            self.all_data = all_data
            self.setWindowTitle("3D Гравитационная система N-тел")
            self.setGeometry(100, 100, 1500, 900)
            self.projection_plane = 'XY'
            self.simulation_history = []  # История симуляции для экспорта
            self.current_time = 0.0  # Текущее время симуляции
            self.init_ui()

        def init_ui(self):
            main_layout = QHBoxLayout()

            # Левая часть - графики
            left_widget = QWidget()
            left_layout = QVBoxLayout(left_widget)

            # Создаем matplotlib canvas с 3D основным и 2D в углу
            self.fig = Figure(figsize=(10, 8))
            self.canvas = FigureCanvas(self.fig)

            # 3D график - основной (занимает большую часть)
            self.ax = self.fig.add_axes([0.05, 0.15, 0.65, 0.80], projection='3d')

            # 2D проекция - маленькая в правом нижнем углу
            self.ax2d = self.fig.add_axes([0.72, 0.10, 0.26, 0.35])

            left_layout.addWidget(self.canvas)

            # Панель управления
            control_layout = QHBoxLayout()
            self.status_label = QLabel("Запуск симуляции...")
            control_layout.addWidget(self.status_label)
            control_layout.addStretch()

            # Выбор плоскости проекции
            projection_label = QLabel("Проекция:")
            control_layout.addWidget(projection_label)

            self.projection_combo = QComboBox()
            self.projection_combo.addItems(['XY', 'XZ', 'YZ'])
            self.projection_combo.currentTextChanged.connect(self.change_projection)
            control_layout.addWidget(self.projection_combo)

            control_layout.addStretch()

            # Кнопка экспорта данных
            self.export_btn = QPushButton("Экспорт данных")
            self.export_btn.clicked.connect(self.export_data)
            self.export_btn.setEnabled(False)
            control_layout.addWidget(self.export_btn)

            # Кнопка паузы/продолжения
            self.pause_btn = QPushButton("Пауза")
            self.pause_btn.clicked.connect(self.toggle_animation)
            self.pause_btn.setMinimumWidth(100)
            control_layout.addWidget(self.pause_btn)

            # Кнопка сброса
            self.reset_btn = QPushButton("Сброс")
            self.reset_btn.clicked.connect(self.reset_simulation)
            self.reset_btn.setToolTip("Сбросить симуляцию к начальным условиям")
            control_layout.addWidget(self.reset_btn)

            left_layout.addLayout(control_layout)

            # Правая часть - таблица с данными
            right_widget = QWidget()
            right_layout = QVBoxLayout(right_widget)

            # Метка времени симуляции
            self.time_label = QLabel("Время: 0.00 с")
            self.time_label.setStyleSheet("font-weight: bold; font-size: 12px;")
            right_layout.addWidget(self.time_label)

            # Таблица координат
            coord_group = QGroupBox("Данные тел")
            coord_layout = QVBoxLayout(coord_group)

            self.coords_table = QTableWidget()
            self.coords_table.setMinimumWidth(320)
            self.coords_table.setStyleSheet("font-family: monospace; font-size: 9px;")
            coord_layout.addWidget(self.coords_table)

            right_layout.addWidget(coord_group)

            # Добавляем виджеты в главный layout
            main_layout.addWidget(left_widget, 3)
            main_layout.addWidget(right_widget, 1)

            self.setLayout(main_layout)

            # Инициализация симуляции
            self.init_simulation()

        def change_projection(self, plane):
            """Изменение плоскости проекции"""
            self.projection_plane = plane

        def init_simulation(self):
            # Начальные условия
            self.n_bodies = len(self.all_data)
            print(f"Количество тел в симуляции: {self.n_bodies}")

            if self.n_bodies == 0:
                self.status_label.setText("Нет тел для симуляции!")
                return

            # Сохраняем как список, а не numpy array
            self.initial_bodies = [body.copy() for body in self.all_data]
            self.bodies = self.initial_bodies.copy()

            # Инициализация истории симуляции
            self.simulation_history = []
            self.current_time = 0.0
            self.save_to_history()  # Сохраняем начальное состояние

            # Настройка таблицы
            self.setup_table()

            # Настройка 3D графика
            self.ax.set_xlabel('X (м)', fontsize=9)
            self.ax.set_ylabel('Y (м)', fontsize=9)
            self.ax.set_zlabel('Z (м)', fontsize=9)
            self.ax.set_title('3D Гравитационная система N-тел', fontsize=10)

            # Настройка 2D графика (маленький)
            self.ax2d.set_xlabel('X', fontsize=8)
            self.ax2d.set_ylabel('Y', fontsize=8)
            self.ax2d.set_title('2D (XY)', fontsize=9)
            self.ax2d.grid(True, alpha=0.3)
            self.ax2d.tick_params(axis='both', labelsize=7)

            # Цвета для каждого тела
            self.colors = ['red', 'blue', 'green', 'orange', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']
            self.trail_length = 1000

            # Массивы для хранения траекторий
            self.trajectories = [[] for _ in range(self.n_bodies)]

            # Создание scatter объектов для тел (3D)
            self.scatters = []
            for i in range(self.n_bodies):
                scatter = self.ax.scatter([self.bodies[i][0]], [self.bodies[i][1]], [self.bodies[i][2]],
                                        c=self.colors[i % len(self.colors)], s=100, label=f'Тело {i+1}')
                self.scatters.append(scatter)

            # Создание линий для траекторий (3D)
            self.lines = []
            for i in range(self.n_bodies):
                line, = self.ax.plot([], [], [], c=self.colors[i % len(self.colors)], alpha=0.7, linewidth=2)
                self.lines.append(line)

            # Создание scatter объектов для 2D проекции
            self.scatters_2d = []
            for i in range(self.n_bodies):
                scatter_2d, = self.ax2d.plot([self.bodies[i][0]], [self.bodies[i][1]], 'o',
                                             c=self.colors[i % len(self.colors)], markersize=6)
                self.scatters_2d.append(scatter_2d)

            # Создание линий для траекторий (2D)
            self.lines_2d = []
            for i in range(self.n_bodies):
                line_2d, = self.ax2d.plot([], [], c=self.colors[i % len(self.colors)], alpha=0.5, linewidth=1)
                self.lines_2d.append(line_2d)

            # Автоматическое определение начальных границ
            bodies_array = np.array(self.bodies)
            positions = bodies_array[:, :3]
            max_coord = np.max(np.abs(positions))
            self.initial_limits = max(1.0, max_coord * 1.5)
            self.current_limits = self.initial_limits
            self.current_limits_2d = self.initial_limits
            self.auto_scale = True

            self.ax.set_xlim(-self.current_limits, self.current_limits)
            self.ax.set_ylim(-self.current_limits, self.current_limits)
            self.ax.set_zlim(-self.current_limits, self.current_limits)

            self.ax2d.set_xlim(-self.current_limits_2d, self.current_limits_2d)
            self.ax2d.set_ylim(-self.current_limits_2d, self.current_limits_2d)

            # Добавление легенды только для 3D
            self.ax.legend(fontsize=7, loc='upper left')

            # Начальное отображение координат
            self.update_coordinates_display()

            # Запуск таймера для анимации
            self.timer = QTimer()
            self.timer.timeout.connect(self.update_animation)
            self.timer.start(50)  # 20 FPS
            self.frame_count = 0
            self.is_running = True

        def setup_table(self):
            """Настройка таблицы для отображения данных тел"""
            # Колонки: Тело, X, Y, Z, Vx, Vy, Vz, Масса
            headers = ['Тело', 'X (м)', 'Y (м)', 'Z (м)', 'Vx (м/с)', 'Vy (м/с)', 'Vz (м/с)', 'Масса (кг)']
            self.coords_table.setColumnCount(len(headers))
            self.coords_table.setRowCount(self.n_bodies)
            self.coords_table.setHorizontalHeaderLabels(headers)

            # Настройка размеров колонок
            header = self.coords_table.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
            for i in range(1, len(headers)):
                header.setSectionResizeMode(i, QHeaderView.Stretch)

            # Заполнение начальными данными
            for i in range(self.n_bodies):
                self.coords_table.setItem(i, 0, QTableWidgetItem(f"Тело {i+1}"))

        def save_to_history(self):
            """Сохранение текущего состояния в историю"""
            snapshot = {
                'time': self.current_time,
                'bodies': [body.copy() for body in self.bodies]
            }
            self.simulation_history.append(snapshot)

        def update_coordinates_display(self):
            """Обновление таблицы координат тел"""
            for i in range(self.n_bodies):
                x, y, z = self.bodies[i][0], self.bodies[i][1], self.bodies[i][2]
                vx, vy, vz = self.bodies[i][3], self.bodies[i][4], self.bodies[i][5]
                mass = self.bodies[i][6]

                self.coords_table.setItem(i, 1, QTableWidgetItem(f"{x:.3e}"))
                self.coords_table.setItem(i, 2, QTableWidgetItem(f"{y:.3e}"))
                self.coords_table.setItem(i, 3, QTableWidgetItem(f"{z:.3e}"))
                self.coords_table.setItem(i, 4, QTableWidgetItem(f"{vx:.3e}"))
                self.coords_table.setItem(i, 5, QTableWidgetItem(f"{vy:.3e}"))
                self.coords_table.setItem(i, 6, QTableWidgetItem(f"{vz:.3e}"))
                self.coords_table.setItem(i, 7, QTableWidgetItem(f"{mass:.3e}"))

        def export_data(self):
            """Экспорт данных симуляции в Excel файл (отдельный лист для каждого тела)"""
            if not self.simulation_history:
                QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
                return

            filename, _ = QFileDialog.getSaveFileName(
                self, "Сохранить данные симуляции",
                f"simulation_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel файлы (*.xlsx);;Все файлы (*)"
            )

            if not filename:
                return

            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

                wb = openpyxl.Workbook()
                # Удаляем стандартный лист
                wb.remove(wb.active)

                # Стили
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Создаем лист для каждого тела
                for body_idx in range(self.n_bodies):
                    sheet_name = f"Тело {body_idx + 1}"
                    ws = wb.create_sheet(title=sheet_name)

                    # Заголовки
                    headers = ['Время (с)', 'Время (читаемое)', 'X (м)', 'Y (м)', 'Z (м)',
                              'Vx (м/с)', 'Vy (м/с)', 'Vz (м/с)', 'Масса (кг)']

                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')

                    # Данные
                    for row_idx, snapshot in enumerate(self.simulation_history, 2):
                        time_sec = snapshot['time']
                        body = snapshot['bodies'][body_idx]

                        # Форматирование времени
                        if time_sec < 60:
                            time_readable = f"{time_sec:.2f} с"
                        elif time_sec < 3600:
                            time_readable = f"{time_sec/60:.2f} мин"
                        elif time_sec < 86400:
                            time_readable = f"{time_sec/3600:.2f} ч"
                        elif time_sec < 31536000:
                            time_readable = f"{time_sec/86400:.2f} дн"
                        else:
                            time_readable = f"{time_sec/31536000:.4f} лет"

                        ws.cell(row=row_idx, column=1, value=time_sec)
                        ws.cell(row=row_idx, column=2, value=time_readable)
                        ws.cell(row=row_idx, column=3, value=body[0])  # X
                        ws.cell(row=row_idx, column=4, value=body[1])  # Y
                        ws.cell(row=row_idx, column=5, value=body[2])  # Z
                        ws.cell(row=row_idx, column=6, value=body[3])  # Vx
                        ws.cell(row=row_idx, column=7, value=body[4])  # Vy
                        ws.cell(row=row_idx, column=8, value=body[5])  # Vz
                        ws.cell(row=row_idx, column=9, value=body[6])  # Mass

                    # Автоширина колонок
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[column].width = min(max_length + 2, 20)

                # Добавляем сводный лист
                ws_summary = wb.create_sheet(title="Сводка", index=0)
                ws_summary.cell(row=1, column=1, value="Информация о симуляции").font = Font(bold=True, size=14)
                ws_summary.cell(row=3, column=1, value="Количество тел:")
                ws_summary.cell(row=3, column=2, value=self.n_bodies)
                ws_summary.cell(row=4, column=1, value="Временных шагов:")
                ws_summary.cell(row=4, column=2, value=len(self.simulation_history))
                ws_summary.cell(row=5, column=1, value="Общее время симуляции:")
                ws_summary.cell(row=5, column=2, value=f"{self.current_time:.6e} с")
                ws_summary.cell(row=6, column=1, value="Шаг времени (dt):")
                ws_summary.cell(row=6, column=2, value=f"{se.dt} с")

                wb.save(filename)

                QMessageBox.information(self, "Успех",
                    f"Данные экспортированы в файл:\n{filename}\n\n"
                    f"Создано листов: {self.n_bodies + 1}\n"
                    f"(Сводка + по листу на каждое тело)\n"
                    f"Временных шагов: {len(self.simulation_history)}")

            except ImportError:
                QMessageBox.warning(self, "Ошибка",
                    "Для экспорта в Excel необходима библиотека openpyxl.\n"
                    "Установите её командой: pip install openpyxl")
            except Exception as e:
                QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить файл:\n{str(e)}")

        def get_projection_coords(self, body):
            """Получение координат для текущей проекции"""
            if self.projection_plane == 'XY':
                return body[0], body[1]
            elif self.projection_plane == 'XZ':
                return body[0], body[2]
            else:  # YZ
                return body[1], body[2]

        def get_projection_labels(self):
            """Получение меток осей для текущей проекции"""
            if self.projection_plane == 'XY':
                return 'X (м)', 'Y (м)'
            elif self.projection_plane == 'XZ':
                return 'X (м)', 'Z (м)'
            else:  # YZ
                return 'Y (м)', 'Z (м)'

        def calcalt(self, l):
            """
            Исправленный интегратор для гравитационной системы N-тел
            """
            G = 6.67430e-11
            n = len(l)
            new_l = [body.copy() for body in l]
    
            # Вычисляем ускорения для всех тел
            accelerations = []
            for i in range(n):
                ax, ay, az = 0, 0, 0
                for j in range(n):
                    if i != j:
                        dx = l[j][0] - l[i][0]
                        dy = l[j][1] - l[i][1]
                        dz = l[j][2] - l[i][2]
                
                        r_sq = dx*dx + dy*dy + dz*dz
                        r = np.sqrt(r_sq)
                
                        # ИСПРАВЛЕНИЕ: правильный расчет силы
                        # F = G * m1 * m2 / r^2
                        # a = F / m1 = G * m2 / r^2
                        force_magnitude = G * l[j][6] / (r_sq * r)  # делим на r^3 для направления
                
                        ax += force_magnitude * dx
                        ay += force_magnitude * dy
                        az += force_magnitude * dz
                
                accelerations.append((ax, ay, az))
    
            # Простой интегратор Эйлера для тестирования
            for i in range(n):
                ax, ay, az = accelerations[i]
        
                # Обновляем скорости
                new_l[i][3] = l[i][3] + ax * se.dt
                new_l[i][4] = l[i][4] + ay * se.dt  
                new_l[i][5] = l[i][5] + az * se.dt
        
                # Обновляем позиции
                new_l[i][0] = l[i][0] + new_l[i][3] * se.dt
                new_l[i][1] = l[i][1] + new_l[i][4] * se.dt
                new_l[i][2] = l[i][2] + new_l[i][5] * se.dt

            if self.frame_count % 100 == 0:
                print(f"Frame {self.frame_count}:")
                print(f"  Body 0 pos: ({l[0][0]:.2e}, {l[0][1]:.2e})")
                print(f"  Body 1 pos: ({l[1][0]:.2e}, {l[1][1]:.2e})")
                print(f"  Distance: {np.sqrt((l[1][0]-l[0][0])**2 + (l[1][1]-l[0][1])**2):.2e}")
    
            return new_l

        def update_animation(self):
            if not self.is_running:
                return

            self.frame_count += 1
            self.current_time += se.dt

            # Обновление позиций
            self.bodies = self.calcalt(self.bodies)

            # Сохраняем в историю каждые 10 кадров для экономии памяти
            if self.frame_count % 10 == 0:
                self.save_to_history()

            # Обновление 3D визуализации
            for i, scatter in enumerate(self.scatters):
                x, y, z = self.bodies[i][0], self.bodies[i][1], self.bodies[i][2]
                scatter._offsets3d = ([x], [y], [z])

                # Обновление траекторий
                self.trajectories[i].append((x, y, z))
                if len(self.trajectories[i]) > self.trail_length:
                    self.trajectories[i].pop(0)

                if len(self.trajectories[i]) > 1:
                    x_traj = [point[0] for point in self.trajectories[i]]
                    y_traj = [point[1] for point in self.trajectories[i]]
                    z_traj = [point[2] for point in self.trajectories[i]]
                    self.lines[i].set_data(x_traj, y_traj)
                    self.lines[i].set_3d_properties(z_traj)

            # Обновление 2D проекции
            xlabel, ylabel = self.get_projection_labels()
            self.ax2d.set_xlabel(xlabel, fontsize=8)
            self.ax2d.set_ylabel(ylabel, fontsize=8)
            self.ax2d.set_title(f'2D ({self.projection_plane})', fontsize=9)

            for i in range(self.n_bodies):
                proj_x, proj_y = self.get_projection_coords(self.bodies[i])
                self.scatters_2d[i].set_data([proj_x], [proj_y])

                # Обновление 2D траекторий
                if len(self.trajectories[i]) > 1:
                    if self.projection_plane == 'XY':
                        traj_x = [point[0] for point in self.trajectories[i]]
                        traj_y = [point[1] for point in self.trajectories[i]]
                    elif self.projection_plane == 'XZ':
                        traj_x = [point[0] for point in self.trajectories[i]]
                        traj_y = [point[2] for point in self.trajectories[i]]
                    else:  # YZ
                        traj_x = [point[1] for point in self.trajectories[i]]
                        traj_y = [point[2] for point in self.trajectories[i]]
                    self.lines_2d[i].set_data(traj_x, traj_y)

            # Автоматическое масштабирование
            if self.auto_scale:
                bodies_array = np.array(self.bodies)
                positions = bodies_array[:, :3]

                # Используем максимальное расстояние от центра
                distances = np.sqrt(np.sum(positions**2, axis=1))
                max_distance = np.max(distances) if len(distances) > 0 else 1.0
                new_limit = max(self.initial_limits * 0.1, max_distance * 1.2)

                # Плавное изменение границ для 3D
                if abs(new_limit - self.current_limits) > self.current_limits * 0.05:
                    self.current_limits = 0.95 * self.current_limits + 0.05 * new_limit
                    self.ax.set_xlim(-self.current_limits, self.current_limits)
                    self.ax.set_ylim(-self.current_limits, self.current_limits)
                    self.ax.set_zlim(-self.current_limits, self.current_limits)

                # Плавное изменение границ для 2D
                if abs(new_limit - self.current_limits_2d) > self.current_limits_2d * 0.05:
                    self.current_limits_2d = 0.95 * self.current_limits_2d + 0.05 * new_limit
                    self.ax2d.set_xlim(-self.current_limits_2d, self.current_limits_2d)
                    self.ax2d.set_ylim(-self.current_limits_2d, self.current_limits_2d)

            # Обновление координат и времени (каждые 5 кадров для производительности)
            if self.frame_count % 5 == 0:
                self.update_coordinates_display()
                self.update_time_display()

            # Обновление статуса
            if self.n_bodies > 0:
                bodies_array = np.array(self.bodies)
                positions = bodies_array[:, :3]
                max_pos = np.max(np.abs(positions))
                self.status_label.setText(f"Кадр: {self.frame_count} | Тела: {self.n_bodies} | Макс: {max_pos:.2e} м")
                self.ax.set_title(f'3D Симуляция | Кадр: {self.frame_count}', fontsize=10)

            # Обновление canvas
            self.canvas.draw_idle()

        def update_time_display(self):
            """Обновление отображения времени симуляции"""
            if self.current_time < 60:
                time_str = f"{self.current_time:.2f} с"
            elif self.current_time < 3600:
                minutes = self.current_time / 60
                time_str = f"{minutes:.2f} мин"
            elif self.current_time < 86400:
                hours = self.current_time / 3600
                time_str = f"{hours:.2f} ч"
            elif self.current_time < 31536000:
                days = self.current_time / 86400
                time_str = f"{days:.2f} дн"
            else:
                years = self.current_time / 31536000
                time_str = f"{years:.4f} лет"

            self.time_label.setText(f"Время: {time_str} ({self.current_time:.2e} с)")

        def toggle_animation(self):
            """Пауза/продолжение симуляции"""
            if self.is_running:
                # Пауза
                self.is_running = False
                self.timer.stop()
                self.save_to_history()
                self.pause_btn.setText("Продолжить")
                self.pause_btn.setStyleSheet("background-color: #90EE90;")  # Светло-зелёный
                self.export_btn.setEnabled(True)
                self.status_label.setText(f"Пауза | Кадр: {self.frame_count} | Записей: {len(self.simulation_history)}")
            else:
                # Продолжение
                self.is_running = True
                self.timer.start(50)
                self.pause_btn.setText("Пауза")
                self.pause_btn.setStyleSheet("")
                self.export_btn.setEnabled(False)
                self.status_label.setText(f"Симуляция возобновлена | Кадр: {self.frame_count}")

        def reset_simulation(self):
            """Сброс симуляции к начальным условиям"""
            # Останавливаем если запущена
            was_running = self.is_running
            if self.is_running:
                self.is_running = False
                self.timer.stop()

            # Подтверждение сброса
            reply = QMessageBox.question(
                self, 'Подтверждение сброса',
                'Сбросить симуляцию к начальным условиям?\nИстория будет очищена.',
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # Восстанавливаем начальные условия
                self.bodies = [body.copy() for body in self.initial_bodies]
                self.frame_count = 0
                self.current_time = 0.0

                # Очищаем историю и траектории
                self.simulation_history = []
                self.save_to_history()
                self.trajectories = [[] for _ in range(self.n_bodies)]

                # Очищаем траектории на графиках
                for i in range(self.n_bodies):
                    self.lines[i].set_data([], [])
                    self.lines[i].set_3d_properties([])
                    self.lines_2d[i].set_data([], [])

                # Обновляем отображение
                self.update_coordinates_display()
                self.update_time_display()
                self.canvas.draw_idle()

                # Сбрасываем кнопки
                self.pause_btn.setText("Старт")
                self.pause_btn.setStyleSheet("background-color: #90EE90;")
                self.export_btn.setEnabled(True)
                self.status_label.setText("Симуляция сброшена | Готов к запуску")
            elif was_running:
                # Если отменили сброс, продолжаем если было запущено
                self.is_running = True
                self.timer.start(50)
    
    # Создаем и показываем диалог
    dialog = AnimationDialog(all_data)
    dialog.exec_()

class ScientificLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setPlaceholderText("Введите число (например: 1.5e10, -3.2, 0.001)")
        
    def get_value(self):
        """Возвращает значение как float или None при ошибке"""
        text = self.text().strip()
        if not text:
            return None
            
        # Заменяем запятые на точки для удобства ввода
        text = text.replace(',', '.')
        
        try:
            value = float(text)
            return value
        except ValueError:
            return None
            
    def set_value(self, value):
        """Устанавливает значение из float"""
        if value is None:
            self.setText("")
        else:
            # Форматируем для отображения
            if abs(value) < 0.001 or abs(value) > 10000:
                self.setText("{:.6e}".format(value))
            else:
                self.setText("{:.10f}".format(value).rstrip('0').rstrip('.'))

class ObjectEditor(QWidget):
    def __init__(self, obj_data=None):
        super().__init__()
        self.obj_data = obj_data or {
            'name': 'Новый объект',
            'x': 0, 'y': 0, 'z': 0,
            'vx': 0, 'vy': 0, 'vz': 0,
            'mass': 1.0
        }
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout()
        
        # Поле имени
        self.name_edit = QLineEdit(self.obj_data['name'])
        layout.addRow("Имя:", self.name_edit)
        
        # Координаты с текстовым вводом
        self.x_edit = ScientificLineEdit()
        self.x_edit.set_value(self.obj_data['x'])
        layout.addRow("X:", self.x_edit)
        
        self.y_edit = ScientificLineEdit()
        self.y_edit.set_value(self.obj_data['y'])
        layout.addRow("Y:", self.y_edit)
        
        self.z_edit = ScientificLineEdit()
        self.z_edit.set_value(self.obj_data['z'])
        layout.addRow("Z:", self.z_edit)
        
        # Скорости с текстовым вводом
        self.vx_edit = ScientificLineEdit()
        self.vx_edit.set_value(self.obj_data['vx'])
        layout.addRow("Vx:", self.vx_edit)
        
        self.vy_edit = ScientificLineEdit()
        self.vy_edit.set_value(self.obj_data['vy'])
        layout.addRow("Vy:", self.vy_edit)
        
        self.vz_edit = ScientificLineEdit()
        self.vz_edit.set_value(self.obj_data['vz'])
        layout.addRow("Vz:", self.vz_edit)
        
        # Масса с текстовым вводом (только положительные значения)
        self.mass_edit = ScientificLineEdit()
        self.mass_edit.set_value(self.obj_data['mass'])
        layout.addRow("Масса:", self.mass_edit)
        
        self.setLayout(layout)

    def get_data(self):
        """Получает данные из формы с проверкой корректности"""
        name = self.name_edit.text().strip()
        if not name:
            name = "Безымянный объект"
            
        # Получаем числовые значения
        x = self.x_edit.get_value() or 0.0
        y = self.y_edit.get_value() or 0.0
        z = self.z_edit.get_value() or 0.0
        vx = self.vx_edit.get_value() or 0.0
        vy = self.vy_edit.get_value() or 0.0
        vz = self.vz_edit.get_value() or 0.0
        
        # Масса должна быть положительной
        mass = self.mass_edit.get_value()
        if mass is None or mass <= 0:
            mass = 1.0  # значение по умолчанию при ошибке
            
        return {
            'name': name,
            'x': x,
            'y': y,
            'z': z,
            'vx': vx,
            'vy': vy,
            'vz': vz,
            'mass': mass
        }

class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Настройки")
        self.setModal(True)
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout()
        
        # Поле для ввода dt как строки
        dt_label = QLabel("Шаг времени (dt):")
        dt_label.setToolTip("Временной шаг симуляции. Может быть очень маленьким (0.001, 0.0001, 1e-5 и т.д.)")
        self.dt_edit = QLineEdit()
        self.dt_edit.setPlaceholderText("Введите значение dt (например: 0.001, 1e-5)")
        
        # Устанавливаем текущее значение
        if hasattr(self.parent, 'settings') and 'dt' in self.parent.settings:
            current_dt = self.parent.settings['dt']
            # Форматируем значение для отображения
            if abs(current_dt) < 0.001 or abs(current_dt) > 10000:
                self.dt_edit.setText("{:.6e}".format(current_dt))
            else:
                self.dt_edit.setText("{:.10f}".format(current_dt).rstrip('0').rstrip('.'))
        
        layout.addRow(dt_label, self.dt_edit)
        
        # Кнопки OK и Отмена
        buttons_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.apply_settings)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(ok_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addRow(buttons_layout)
        
        self.setLayout(layout)

    def parse_float_value(self, text):
        """Парсит строку в float с обработкой ошибок"""
        text = text.strip()
        if not text:
            return None, "Поле не может быть пустым"
            
        # Заменяем запятые на точки для удобства ввода
        text = text.replace(',', '.')
        
        try:
            value = float(text)
            if value <= 0:
                return None, "Значение должно быть положительным"
            if value > 2592000:  # 30 дней в секундах
                return None, "Значение слишком большое (максимум 2592000)"
            return value, None
        except ValueError:
            return None, "Некорректный формат числа. Используйте: 0.001, 1e-5, 0.00001 и т.д."

    def apply_settings(self):
        """Применяет настройки при нажатии OK"""
        dt_text = self.dt_edit.text().strip()
        dt_value, error = self.parse_float_value(dt_text)
        
        if error:
            QMessageBox.warning(self, "Ошибка ввода", error)
            return
            
        # Сохраняем настройки
        if hasattr(self.parent, 'settings'):
            self.parent.settings['dt'] = dt_value
            
        # Сохраняем настройки
        if self.parent.save_current_state(silent=True):
            self.accept()
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось сохранить настройки")

    def get_settings(self):
        """Возвращает текущие настройки (для обратной совместимости)"""
        dt_text = self.dt_edit.text().strip()
        dt_value, error = self.parse_float_value(dt_text)
        if error:
            return {'dt': 0.1}  # Значение по умолчанию при ошибке
        return {'dt': dt_value}

class SaveDialog(QDialog):
    def __init__(self, parent=None, current_name=""):
        super().__init__(parent)
        self.setWindowTitle("Сохранение")
        self.setModal(True)
        self.init_ui(current_name)

    def init_ui(self, current_name):
        layout = QVBoxLayout()
        form_layout = QFormLayout()
        
        self.name_edit = QLineEdit(current_name)
        form_layout.addRow("Название сохранения:", self.name_edit)
        
        layout.addLayout(form_layout)
        
        buttons_layout = QHBoxLayout()
        save_btn = QPushButton("Сохранить")
        save_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)
        
        self.setLayout(layout)

    def get_save_name(self):
        return self.name_edit.text().strip()

class SaveManager:
    @staticmethod
    def get_saves_list():
        saves = []
        if not os.path.exists('saves'):
            os.makedirs('saves')
            
        for file in os.listdir('saves'):
            if file.endswith('.json'):
                try:
                    with open(os.path.join('saves', file), 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    saves.append({
                        'name': data.get('name', 'Безымянное сохранение'),
                        'date': data.get('date', 'Неизвестно'),
                        'filename': file
                    })
                except Exception as e:
                    print(f"Ошибка загрузки файла {file}: {e}")
        return saves

    @staticmethod
    def load_save(filename):
        try:
            with open(os.path.join('saves', filename), 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Ошибка загрузки сохранения {filename}: {e}")
            return None

    @staticmethod
    def create_save(name, data):
        if not name:
            name = "Безымянное сохранение"
            
        # Создаем безопасное имя файла
        safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{safe_name}_{timestamp}.json"
        
        save_data = {
            'name': name,
            'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data': data
        }
        
        try:
            with open(os.path.join('saves', filename), 'w', encoding='utf-8') as f:
                json.dump(save_data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"Ошибка сохранения: {e}")
            return False

    @staticmethod
    def delete_save(filename):
        try:
            os.remove(os.path.join('saves', filename))
            return True
        except Exception as e:
            print(f"Ошибка удаления файла {filename}: {e}")
            return False

    @staticmethod
    def update_save(filename, data, silent=False):
        try:
            with open(os.path.join('saves', filename), 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
                
            existing_data['data'] = data
            existing_data['date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            with open(os.path.join('saves', filename), 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, indent=2, ensure_ascii=False)
                
            if not silent:
                print("Сохранение успешно обновлено")
            return True
        except Exception as e:
            print(f"Ошибка обновления сохранения {filename}: {e}")
            return False

class MainMenu(QMainWindow):
    def __init__(self):
        super().__init__()
        self.saves_visible = False
        self.current_save_filename = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Главное меню")
        self.setGeometry(100, 100, 500, 400)
        
        central_widget = QWidget()
        layout = QVBoxLayout()
        
        # Кнопки управления сохранениями
        saves_management_layout = QHBoxLayout()
        self.save_btn = QPushButton("Выбрать сохранение")
        self.save_btn.clicked.connect(self.toggle_saves_list)
        saves_management_layout.addWidget(self.save_btn)
        
        new_save_btn = QPushButton("Новое сохранение")
        new_save_btn.clicked.connect(self.create_new_save)
        saves_management_layout.addWidget(new_save_btn)
        
        layout.addLayout(saves_management_layout)
        
        # Список сохранений (изначально скрыт)
        self.saves_list = QListWidget()
        self.saves_list.itemDoubleClicked.connect(self.select_save)
        self.saves_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.saves_list.customContextMenuRequested.connect(self.show_saves_context_menu)
        self.saves_list.hide()
        layout.addWidget(self.saves_list)
        
        # Загрузка списка сохранений
        self.load_saves()
        
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def toggle_saves_list(self):
        self.saves_visible = not self.saves_visible
        self.saves_list.setVisible(self.saves_visible)
        if self.saves_visible:
            self.load_saves()

    def load_saves(self):
        self.saves_list.clear()
        saves = SaveManager.get_saves_list()
        for save in saves:
            item = QListWidgetItem(f"{save['name']} - {save['date']}")
            item.setData(Qt.UserRole, save['filename'])
            self.saves_list.addItem(item)

    def show_saves_context_menu(self, position):
        item = self.saves_list.itemAt(position)
        if item:
            menu = QMenu(self)
            load_action = QAction("Загрузить", self)
            load_action.triggered.connect(lambda: self.select_save(item))
            menu.addAction(load_action)
            
            delete_action = QAction("Удалить", self)
            delete_action.triggered.connect(lambda: self.delete_save(item))
            menu.addAction(delete_action)
            
            menu.exec_(self.saves_list.mapToGlobal(position))

    def delete_save(self, item):
        filename = item.data(Qt.UserRole)
        save_name = item.text().split(' - ')[0]
        
        reply = QMessageBox.question(self, 'Подтверждение удаления',
                                    f'Вы уверены, что хотите удалить сохранение "{save_name}"?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            if SaveManager.delete_save(filename):
                self.load_saves()
                QMessageBox.information(self, "Успех", "Сохранение удалено")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить сохранение")

    def select_save(self, item):
        global SELECTED_SAVE
        filename = item.data(Qt.UserRole)
        SELECTED_SAVE = SaveManager.load_save(filename)
        
        if SELECTED_SAVE:
            self.current_save_filename = filename
            self.open_simulation_menu()
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось загрузить сохранение")

    def create_new_save(self):
        dialog = SaveDialog(self)
        if dialog.exec_():
            save_name = dialog.get_save_name()
            if not save_name:
                save_name = "Новое сохранение"
                
            # Создаем пустые данные для нового сохранения
            new_save_data = {
                'objects': [],
                'settings': {
                    'dt': 0.1
                }
            }
            
            if SaveManager.create_save(save_name, new_save_data):
                self.load_saves()
                QMessageBox.information(self, "Успех", "Новое сохранение создано")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось создать сохранение")

    def open_simulation_menu(self):
        self.simulation_menu = SimulationMenu(self.current_save_filename)
        self.simulation_menu.show()
        self.hide()

class SimulationMenu(QMainWindow):
    def __init__(self, save_filename):
        super().__init__()
        self.save_filename = save_filename
        self.objects = SELECTED_SAVE.get('data', {}).get('objects', [])
        self.settings = SELECTED_SAVE.get('data', {}).get('settings', {})
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"Меню симуляции - {SELECTED_SAVE.get('name', 'Безымянное сохранение')}")
        self.setGeometry(100, 100, 900, 700)
        
        central_widget = QWidget()
        main_layout = QHBoxLayout()
        
        # Левая панель - список объектов
        left_panel = QVBoxLayout()
        left_panel.addWidget(QLabel("Объекты:"))
        
        self.objects_list = QListWidget()
        self.objects_list.itemClicked.connect(self.show_object_editor)
        left_panel.addWidget(self.objects_list)
        
        # Кнопки для объектов
        obj_buttons_layout = QHBoxLayout()
        add_obj_btn = QPushButton("Добавить объект")
        add_obj_btn.clicked.connect(self.add_object)
        remove_obj_btn = QPushButton("Удалить объект")
        remove_obj_btn.clicked.connect(self.remove_object)
        obj_buttons_layout.addWidget(add_obj_btn)
        obj_buttons_layout.addWidget(remove_obj_btn)
        left_panel.addLayout(obj_buttons_layout)
        
        # Область редактирования объекта
        self.object_editor_area = QScrollArea()
        self.object_editor_area.setWidgetResizable(True)
        left_panel.addWidget(self.object_editor_area)
        
        # Правая панель - кнопки управления
        right_panel = QVBoxLayout()
        
        # Информация о текущих настройках
        self.settings_info = QLabel(f"Текущий dt: {self.settings.get('dt', 0.1)}")
        self.settings_info.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 5px; border: 1px solid #ccc; }")
        right_panel.addWidget(self.settings_info)
        
        # Кнопка настроек
        settings_btn = QPushButton("Настройки (dt)")
        settings_btn.clicked.connect(self.open_settings)
        settings_btn.setToolTip("Настройка временного шага симуляции")
        right_panel.addWidget(settings_btn)
        
        # Кнопка сохранения
        save_btn = QPushButton("Сохранить")
        save_btn.clicked.connect(self.save_current_state)
        right_panel.addWidget(save_btn)
        
        # Кнопка сохранения как
        save_as_btn = QPushButton("Сохранить как...")
        save_as_btn.clicked.connect(self.save_as)
        right_panel.addWidget(save_as_btn)
        
        # Кнопка старта симуляции
        start_btn = QPushButton("Запустить симуляцию")
        start_btn.clicked.connect(self.start_simulation)
        start_btn.setMinimumHeight(80)
        right_panel.addWidget(start_btn)
        
        right_panel.addStretch()
        
        # Верхняя панель - навигация
        top_panel = QHBoxLayout()
        back_btn = QPushButton("Назад в меню")
        back_btn.clicked.connect(self.back_to_main)
        top_panel.addWidget(back_btn)
        top_panel.addStretch()
        
        # Основной layout
        main_layout.addLayout(left_panel, 3)
        main_layout.addLayout(right_panel, 1)
        
        central_layout = QVBoxLayout()
        central_layout.addLayout(top_panel)
        central_layout.addLayout(main_layout)
        
        central_widget.setLayout(central_layout)
        self.setCentralWidget(central_widget)
        
        self.load_objects()

    def load_objects(self):
        self.objects_list.clear()
        for obc in self.objects:
            self.objects_list.addItem(obc['name'])

    def show_object_editor(self, item):
        index = self.objects_list.row(item)
        editor = ObjectEditor(self.objects[index])
        self.object_editor_area.setWidget(editor)
        self.current_editor = editor
        self.current_editor_index = index

    def add_object(self):
        new_obj = {
            'name': f'Объект {len(self.objects) + 1}',
            'x': 0, 'y': 0, 'z': 0,
            'vx': 0, 'vy': 0, 'vz': 0,
            'mass': 1.0
        }
        self.objects.append(new_obj)
        self.load_objects()
        self.save_current_state(silent=True)

    def remove_object(self):
        current_row = self.objects_list.currentRow()
        if current_row >= 0:
            self.objects.pop(current_row)
            self.load_objects()
            self.object_editor_area.setWidget(None)
            self.save_current_state(silent=True)

    def save_current_state(self, silent=False):
        # Обновляем данные текущего редактируемого объекта
        if hasattr(self, 'current_editor') and hasattr(self, 'current_editor_index'):
            self.objects[self.current_editor_index] = self.current_editor.get_data()
            
        # Сохраняем все данные
        save_data = {
            'objects': self.objects,
            'settings': self.settings
        }
        
        if SaveManager.update_save(self.save_filename, save_data, silent=silent):
            if not silent:
                QMessageBox.information(self, "Успех", "Сохранение обновлено")
                
            # Обновляем глобальную переменную
            global SELECTED_SAVE
            SELECTED_SAVE = SaveManager.load_save(self.save_filename)
            
            # Обновляем отображение настроек
            self.update_settings_display()
            return True
        else:
            if not silent:
                QMessageBox.warning(self, "Ошибка", "Не удалось сохранить изменения")
            return False

    def save_as(self):
        current_name = SELECTED_SAVE.get('name', '')
        dialog = SaveDialog(self, current_name)
        if dialog.exec_():
            save_name = dialog.get_save_name()
            if not save_name:
                save_name = "Копия сохранения"
                
            save_data = {
                'objects': self.objects,
                'settings': self.settings
            }
            
            if SaveManager.create_save(save_name, save_data):
                QMessageBox.information(self, "Успех", "Сохранение создано")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось создать сохранение")

    def open_settings(self):
        dialog = SettingsDialog(self)
        if dialog.exec_():
            # Настройки уже сохранены в apply_settings
            pass

    def update_settings_display(self):
        dt_value = self.settings.get('dt', 0.1)
        if abs(dt_value) < 0.001 or abs(dt_value) > 10000:
            dt_display = "{:.6e}".format(dt_value)
        else:
            dt_display = "{:.10f}".format(dt_value).rstrip('0').rstrip('.')
        self.settings_info.setText(f"Текущий dt: {dt_display}")

    def start_simulation(self):
        global all_data
        all_data = []
        
        # Сохраняем текущее состояние перед запуском
        if not self.save_current_state(silent=True):
            QMessageBox.warning(self, "Ошибка", "Не удалось сохранить состояние перед запуском симуляции")
            return
            
        # Преобразуем объекты в формат для симуляции
        for obj in self.objects:
            # [x, y, z, vx, vy, vz, mass]
            all_data.append([
                obj['x'], obj['y'], obj['z'],
                obj['vx'], obj['vy'], obj['vz'],
                obj['mass']
            ])
        
        # Устанавливаем dt в глобальную переменную
        se.dt = self.settings.get('dt', 0.1)
        
        # Запускаем симуляцию
        animation3dyes(all_data)

    def back_to_main(self):
        self.main_menu = MainMenu()
        self.main_menu.show()
        self.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_menu = MainMenu()
    main_menu.show()
    sys.exit(app.exec_())